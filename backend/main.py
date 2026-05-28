from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, WebSocket, WebSocketDisconnect
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from sqlalchemy import case, func, or_
from typing import List, Optional
from datetime import datetime, timedelta
import os
from pathlib import Path
import asyncio
import pandas as pd
from io import BytesIO
from fastapi.responses import StreamingResponse

import models, schemas, database, auth_service, export_service, email_service

# Initialize Database
models.Base.metadata.create_all(bind=database.engine)

app = FastAPI(title="ISP Fault Management System")

# CORS
DEFAULT_CORS_ORIGINS = "http://127.0.0.1:5173,http://localhost:5173"
ALLOWED_ORIGINS = [
    origin.strip()
    for origin in os.getenv("CORS_ORIGINS", DEFAULT_CORS_ORIGINS).split(",")
    if origin.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

VALID_ROLES = {"Admin", "Engineer", "Viewer"}
VALID_STATUSES = {"Open", "Assigned", "Investigating", "Waiting for ISP", "Resolved", "Closed"}
ACTIVE_STATUSES = {"Open", "Assigned", "Investigating", "Waiting for ISP"}
VALID_SEVERITIES = {"Minor", "Major", "Critical"}
ALLOWED_UPLOAD_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".pdf"}
ALLOWED_UPLOAD_TYPES = {"image/jpeg", "image/png", "image/webp", "application/pdf"}
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))

def _migrate_sqlite_schema():
    if database.engine.dialect.name != "sqlite":
        return

    fault_columns = {
        "ticket_number": "VARCHAR",
        "updated_at": "DATETIME",
        "closed_at": "DATETIME",
        "assigned_to": "VARCHAR",
        "resolution_note": "TEXT",
    }

    with database.engine.begin() as conn:
        existing = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(fault_logs)").fetchall()}
        for column, definition in fault_columns.items():
            if column not in existing:
                conn.exec_driver_sql(f"ALTER TABLE fault_logs ADD COLUMN {column} {definition}")
        notification_existing = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(notification_rules)").fetchall()}
        if "stakeholder_name" not in notification_existing:
            conn.exec_driver_sql("ALTER TABLE notification_rules ADD COLUMN stakeholder_name VARCHAR")
        conn.exec_driver_sql(
            "CREATE UNIQUE INDEX IF NOT EXISTS ix_fault_logs_ticket_number "
            "ON fault_logs (ticket_number) WHERE ticket_number IS NOT NULL"
        )

_migrate_sqlite_schema()

@app.on_event("startup")
def startup_event():
    db = next(database.get_db())
    try:
        _backfill_ticket_numbers(db)

        # Create the first admin only when explicit bootstrap credentials are provided.
        if db.query(models.User).count() == 0:
            bootstrap_username = os.getenv("INITIAL_ADMIN_USERNAME")
            bootstrap_password = os.getenv("INITIAL_ADMIN_PASSWORD")
            bootstrap_email = os.getenv("INITIAL_ADMIN_EMAIL", "admin@bcc.gov.zw")
            if not bootstrap_username or not bootstrap_password:
                print(
                    "No users exist. Set INITIAL_ADMIN_USERNAME and INITIAL_ADMIN_PASSWORD "
                    "to bootstrap the first administrator."
                )
                return
            if len(bootstrap_password) < 8 and not _weak_local_passwords_allowed():
                print("INITIAL_ADMIN_PASSWORD must be at least 8 characters.")
                return
            hashed_pwd = auth_service.get_password_hash(bootstrap_password)
            new_admin = models.User(
                username=bootstrap_username,
                email=bootstrap_email,
                password_hash=hashed_pwd, 
                role="Admin"
            )
            db.add(new_admin)
            db.commit()
            print(f"Initial admin user created: {bootstrap_username}")
    except Exception as e:
        print(f"Error during startup: {e}")
    finally:
        db.close()

# Static Files
if not os.path.exists("uploads"):
    os.makedirs("uploads")
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

# Auth
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# WebSocket Manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            await connection.send_json(message)

manager = ConnectionManager()

# Dependency
async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(database.get_db)):
    credentials_exception = HTTPException(
        status_code=401,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = auth_service.jwt.decode(token, auth_service.SECRET_KEY, algorithms=[auth_service.ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except auth_service.JWTError:
        raise credentials_exception
        
    user = db.query(models.User).filter(models.User.username == username).first()
    if user is None:
        raise credentials_exception
    return user

# Auth Routes
@app.post("/token", response_model=schemas.Token)
def login_for_access_token(db: Session = Depends(database.get_db), form_data: OAuth2PasswordRequestForm = Depends()):
    print(f"Login attempt for user: {form_data.username}")
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    if not user:
        print(f"User {form_data.username} not found in database.")
    elif not auth_service.verify_password(form_data.password, user.password_hash):
        print(f"Password mismatch for user {form_data.username}.")
    
    if not user or not auth_service.verify_password(form_data.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    access_token = auth_service.create_access_token(data={"sub": user.username})
    return {"access_token": access_token, "token_type": "bearer"}

def _require_admin(current_user: models.User):
    if getattr(current_user, "role", None) != "Admin":
        raise HTTPException(status_code=403, detail="Admin privileges required")

def _require_roles(current_user: models.User, allowed_roles: set[str]):
    if getattr(current_user, "role", None) not in allowed_roles:
        raise HTTPException(status_code=403, detail="Insufficient privileges")

def _normalize_role(role: str) -> str:
    role_map = {r.lower(): r for r in VALID_ROLES}
    normalized = role_map.get((role or "").strip().lower())
    if not normalized:
        raise HTTPException(status_code=400, detail="role must be Admin, Engineer, or Viewer")
    return normalized

def _weak_local_passwords_allowed() -> bool:
    enabled = os.getenv("ALLOW_WEAK_LOCAL_PASSWORDS", "").strip().lower() in {"1", "true", "yes"}
    return enabled and database.engine.dialect.name == "sqlite"

def _validate_password(password: str, allow_weak_local: bool = False):
    if not password:
        raise HTTPException(status_code=400, detail="password is required")
    if len(password) < 8 and not (allow_weak_local and _weak_local_passwords_allowed()):
        raise HTTPException(status_code=400, detail="password must be at least 8 characters")

def _validate_fault_values(severity: Optional[str] = None, status: Optional[str] = None):
    if severity is not None and severity not in VALID_SEVERITIES:
        raise HTTPException(status_code=400, detail="severity must be Minor, Major, or Critical")
    if status is not None and status not in VALID_STATUSES:
        raise HTTPException(
            status_code=400,
            detail="status must be Open, Assigned, Investigating, Waiting for ISP, Resolved, or Closed",
        )

def _generate_ticket_number(db: Session, created_at: Optional[datetime] = None) -> str:
    year = (created_at or datetime.now()).year
    prefix = f"BCC-NET-{year}-"
    existing_count = db.query(models.FaultLog).filter(models.FaultLog.ticket_number.like(f"{prefix}%")).count()
    next_number = existing_count + 1
    while True:
        ticket = f"{prefix}{next_number:04d}"
        exists = db.query(models.FaultLog).filter(models.FaultLog.ticket_number == ticket).first()
        if not exists:
            return ticket
        next_number += 1

def _backfill_ticket_numbers(db: Session):
    faults = (
        db.query(models.FaultLog)
        .filter(or_(models.FaultLog.ticket_number == None, models.FaultLog.ticket_number == ""))  # noqa: E711
        .order_by(models.FaultLog.created_at.asc(), models.FaultLog.id.asc())
        .all()
    )
    for fault in faults:
        fault.ticket_number = _generate_ticket_number(db, fault.created_at)
        if not fault.updated_at:
            fault.updated_at = fault.created_at or datetime.now()
    if faults:
        db.commit()

def _record_timeline(db: Session, fault_id: int, action: str, details: str, actor: str):
    db.add(models.FaultTimeline(fault_id=fault_id, action=action, details=details, actor=actor))

def _severity_to_tier(severity: str) -> str:
    s = (severity or "").strip().lower()
    if s == "critical":
        return "CRITICAL"
    if s == "major":
        return "WARNING"
    return "INFO"

def _build_outage_email(fault: models.FaultLog) -> tuple[str, str, str]:
    tier = _severity_to_tier(fault.severity)
    subject = f"{tier}: Network Outage Detected - {fault.location} ({fault.isp_name})"
    detected = fault.created_at.strftime("%d %b %Y, %I:%M %p") if fault.created_at else "Unknown"
    body = (
        "Dear Stakeholder,\n\n"
        "BCC NIMS has automatically recorded a network disruption:\n"
        f"  Provider  : {fault.isp_name}\n"
        f"  Location  : {fault.location}\n"
        f"  Severity  : {fault.severity}\n"
        f"  Detected  : {detected}\n"
        f"  Status    : {fault.status}\n"
        f"  Ref #     : {fault.ticket_number or fault.id}\n\n"
        "IT Operations has been notified and is responding.\n\n"
        "-- BCC NIMS (Automated Alert)\n"
    )
    return tier, subject, body

@app.get("/users/me", response_model=schemas.UserOut)
def read_current_user(current_user: models.User = Depends(get_current_user)):
    return current_user

@app.get("/users/assignees", response_model=List[schemas.UserOut])
def list_assignees(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_roles(current_user, {"Admin", "Engineer"})
    return (
        db.query(models.User)
        .filter(models.User.role.in_(["Admin", "Engineer"]))
        .order_by(models.User.username.asc())
        .all()
    )

@app.get("/users/", response_model=List[schemas.UserOut])
def list_users(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_admin(current_user)
    return db.query(models.User).order_by(models.User.username.asc()).all()

@app.post("/users/", response_model=schemas.UserOut)
def create_user(
    user: schemas.UserCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_admin(current_user)
    role = _normalize_role(user.role)
    _validate_password(user.password)
    db_user = db.query(models.User).filter(models.User.username == user.username).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    db_email = db.query(models.User).filter(models.User.email == user.email).first()
    if db_email:
        raise HTTPException(status_code=400, detail="Email already registered")
    hashed_password = auth_service.get_password_hash(user.password)
    new_user = models.User(username=user.username, email=user.email, password_hash=hashed_password, role=role)
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@app.patch("/users/{user_id}", response_model=schemas.UserOut)
def update_user(
    user_id: int,
    update_data: schemas.UserUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_admin(current_user)
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")

    data = update_data.model_dump(exclude_unset=True)
    if "role" in data and data["role"] is not None:
        db_user.role = _normalize_role(data["role"])
    if "email" in data and data["email"] is not None:
        db_user.email = data["email"]
    if "password" in data and data["password"]:
        _validate_password(data["password"])
        db_user.password_hash = auth_service.get_password_hash(data["password"])

    db.commit()
    db.refresh(db_user)
    return db_user

# Notification Rules (Upper Management Alerts)
@app.get("/notification-rules/", response_model=List[schemas.NotificationRuleOut])
def list_notification_rules(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_admin(current_user)
    return db.query(models.NotificationRule).order_by(models.NotificationRule.created_at.desc()).all()

@app.post("/notification-rules/", response_model=schemas.NotificationRuleOut)
def create_notification_rule(
    rule: schemas.NotificationRuleCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_admin(current_user)
    tier = (rule.tier or "").strip().upper()
    if tier not in {"INFO", "WARNING", "CRITICAL"}:
        raise HTTPException(status_code=400, detail="tier must be INFO, WARNING, or CRITICAL")
    stakeholder_name = (rule.stakeholder_name or "").strip() or None
    db_rule = models.NotificationRule(
        stakeholder_name=stakeholder_name,
        tier=tier,
        email=rule.email,
        is_enabled=rule.is_enabled,
    )
    db.add(db_rule)
    db.commit()
    db.refresh(db_rule)
    return db_rule

@app.patch("/notification-rules/{rule_id}", response_model=schemas.NotificationRuleOut)
def update_notification_rule(
    rule_id: int,
    update_data: schemas.NotificationRuleUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_admin(current_user)
    db_rule = db.query(models.NotificationRule).filter(models.NotificationRule.id == rule_id).first()
    if not db_rule:
        raise HTTPException(status_code=404, detail="Notification rule not found")

    data = update_data.dict(exclude_unset=True)
    if "tier" in data:
        tier = (data["tier"] or "").strip().upper()
        if tier not in {"INFO", "WARNING", "CRITICAL"}:
            raise HTTPException(status_code=400, detail="tier must be INFO, WARNING, or CRITICAL")
        data["tier"] = tier
    if "stakeholder_name" in data and data["stakeholder_name"] is not None:
        data["stakeholder_name"] = data["stakeholder_name"].strip() or None

    for k, v in data.items():
        setattr(db_rule, k, v)

    db.commit()
    db.refresh(db_rule)
    return db_rule

@app.delete("/notification-rules/{rule_id}")
def delete_notification_rule(
    rule_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_admin(current_user)
    db_rule = db.query(models.NotificationRule).filter(models.NotificationRule.id == rule_id).first()
    if not db_rule:
        raise HTTPException(status_code=404, detail="Notification rule not found")
    db.delete(db_rule)
    db.commit()
    return {"message": "Notification rule deleted"}

# Fault Logs Routes
@app.post("/faults/", response_model=schemas.FaultLog)
async def create_fault(
    fault: schemas.FaultLogCreate, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_roles(current_user, {"Admin", "Engineer"})
    _validate_fault_values(severity=fault.severity)

    fault_data = fault.model_dump(exclude={"logged_by"})
    assigned_to = (fault_data.pop("assigned_to", None) or "").strip() or None
    if assigned_to:
        assignee = db.query(models.User).filter(models.User.username == assigned_to).first()
        if not assignee or assignee.role not in {"Admin", "Engineer"}:
            raise HTTPException(status_code=400, detail="assigned_to must be an Admin or Engineer username")

    now = datetime.now()
    db_fault = models.FaultLog(
        **fault_data,
        assigned_to=assigned_to,
        status="Assigned" if assigned_to else "Open",
        ticket_number=_generate_ticket_number(db, now),
        logged_by=current_user.username,
        created_at=now,
        updated_at=now,
    )
    db.add(db_fault)
    db.flush()
    _record_timeline(
        db,
        db_fault.id,
        "Created",
        f"{db_fault.ticket_number} logged for {db_fault.location} on {db_fault.isp_name}.",
        current_user.username,
    )
    if assigned_to:
        _record_timeline(db, db_fault.id, "Assigned", f"Assigned to {assigned_to}.", current_user.username)
    db.commit()
    db.refresh(db_fault)

    # Dispatch tiered emails (configured via /notification-rules/)
    tier, subject, body = _build_outage_email(db_fault)
    recipients = [
        r.email for r in db.query(models.NotificationRule)
        .filter(models.NotificationRule.tier == tier, models.NotificationRule.is_enabled == True)  # noqa: E712
        .all()
    ]
    for recipient in recipients:
        asyncio.create_task(email_service.EmailService.send_email(subject=subject, recipient=recipient, body=body))
    
    # Broadcast new fault
    alert = {
        "event": "new_fault",
        "isp": db_fault.isp_name,
        "severity": db_fault.severity,
        "id": db_fault.id,
        "ticket": db_fault.ticket_number,
    }
    await manager.broadcast(alert)
    return db_fault

@app.get("/faults/", response_model=List[schemas.FaultLog])
def read_faults(
    period: str = Query("all", enum=["day", "week", "month", "all"]),
    isp: Optional[str] = None,
    severity: Optional[str] = None,
    location: Optional[str] = None,
    status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    query = db.query(models.FaultLog)
    
    if period == "day":
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        query = query.filter(models.FaultLog.created_at >= today)
    elif period == "week":
        week_ago = datetime.now() - timedelta(days=7)
        query = query.filter(models.FaultLog.created_at >= week_ago)
    elif period == "month":
        month_ago = datetime.now() - timedelta(days=30)
        query = query.filter(models.FaultLog.created_at >= month_ago)

    if isp:
        query = query.filter(models.FaultLog.isp_name == isp)
    if severity:
        _validate_fault_values(severity=severity)
        query = query.filter(models.FaultLog.severity == severity)
    if location:
        query = query.filter(models.FaultLog.location.ilike(f"%{location}%"))
    if status:
        _validate_fault_values(status=status)
        query = query.filter(models.FaultLog.status == status)
    if assigned_to:
        query = query.filter(models.FaultLog.assigned_to == assigned_to)
    if date_from:
        query = query.filter(models.FaultLog.created_at >= date_from)
    if date_to:
        query = query.filter(models.FaultLog.created_at <= date_to)
    if search:
        needle = f"%{search.strip()}%"
        search_filters = [
            models.FaultLog.ticket_number.ilike(needle),
            models.FaultLog.isp_name.ilike(needle),
            models.FaultLog.location.ilike(needle),
            models.FaultLog.fault_type.ilike(needle),
            models.FaultLog.description.ilike(needle),
        ]
        if search.strip().isdigit():
            search_filters.append(models.FaultLog.id == int(search.strip()))
        query = query.filter(or_(*search_filters))

    severity_order = case(
        (models.FaultLog.severity == "Critical", 0),
        (models.FaultLog.severity == "Major", 1),
        else_=2,
    )
    status_order = case(
        (models.FaultLog.status.in_(["Open", "Assigned", "Investigating", "Waiting for ISP"]), 0),
        else_=1,
    )
    return query.order_by(status_order, severity_order, models.FaultLog.created_at.desc()).all()

@app.patch("/faults/{fault_id}", response_model=schemas.FaultLog)
async def update_fault(
    fault_id: int, 
    update_data: schemas.FaultLogUpdate, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_roles(current_user, {"Admin", "Engineer"})
    db_fault = db.query(models.FaultLog).filter(models.FaultLog.id == fault_id).first()
    if not db_fault:
        raise HTTPException(status_code=404, detail="Fault not found")

    data = update_data.model_dump(exclude_unset=True)
    if "status" in data:
        _validate_fault_values(status=data["status"])
        if data["status"] in {"Resolved", "Closed"} and not (data.get("resolution_note") or db_fault.resolution_note):
            raise HTTPException(status_code=400, detail="resolution_note is required before resolving or closing a fault")
    if "assigned_to" in data:
        assigned_to = (data["assigned_to"] or "").strip() or None
        if assigned_to:
            assignee = db.query(models.User).filter(models.User.username == assigned_to).first()
            if not assignee or assignee.role not in {"Admin", "Engineer"}:
                raise HTTPException(status_code=400, detail="assigned_to must be an Admin or Engineer username")
        data["assigned_to"] = assigned_to

    for key, value in data.items():
        old_value = getattr(db_fault, key, None)
        if old_value == value:
            continue
        setattr(db_fault, key, value)
        if key == "status":
            if value == "Resolved":
                db_fault.resolved_at = db_fault.resolved_at or datetime.now()
                db_fault.closed_at = None
            elif value == "Closed":
                db_fault.resolved_at = db_fault.resolved_at or datetime.now()
                db_fault.closed_at = datetime.now()
            else:
                db_fault.resolved_at = None
                db_fault.closed_at = None
            _record_timeline(db, db_fault.id, "Status Changed", f"{old_value or 'Unknown'} -> {value}.", current_user.username)
        elif key == "assigned_to":
            _record_timeline(
                db,
                db_fault.id,
                "Assigned",
                f"Assigned to {value or 'Unassigned'}.",
                current_user.username,
            )
        elif key == "resolution_note":
            _record_timeline(db, db_fault.id, "Resolution Note", value, current_user.username)
        elif key == "is_sla_breach":
            _record_timeline(db, db_fault.id, "SLA Updated", f"SLA breach set to {value}.", current_user.username)

    db_fault.updated_at = datetime.now()
    
    db.commit()
    db.refresh(db_fault)
    
    if data.get("status") == "Resolved":
        await manager.broadcast({"event": "fault_resolved", "id": fault_id, "ticket": db_fault.ticket_number, "isp": db_fault.isp_name})
    elif data.get("status") == "Closed":
        await manager.broadcast({"event": "fault_closed", "id": fault_id, "ticket": db_fault.ticket_number, "isp": db_fault.isp_name})
    else:
        await manager.broadcast({"event": "fault_updated", "id": fault_id, "ticket": db_fault.ticket_number, "isp": db_fault.isp_name})
        
    return db_fault

@app.delete("/faults/{fault_id}")
async def delete_fault(
    fault_id: int, 
    delete_data: schemas.FaultLogDelete,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_admin(current_user)
    # Verify password
    if not auth_service.verify_password(delete_data.password, current_user.password_hash):
        raise HTTPException(status_code=401, detail="Incorrect password for deletion")
        
    db_fault = db.query(models.FaultLog).filter(models.FaultLog.id == fault_id).first()
    if not db_fault:
        raise HTTPException(status_code=404, detail="Fault not found")

    db.query(models.FaultComment).filter(models.FaultComment.fault_id == fault_id).delete()
    db.query(models.FaultTimeline).filter(models.FaultTimeline.fault_id == fault_id).delete()
    db.delete(db_fault)
    db.commit()
    
    await manager.broadcast({"event": "fault_deleted", "id": fault_id, "ticket": db_fault.ticket_number})
    return {"message": "Fault deleted successfully"}

# File Upload
@app.post("/upload/{fault_id}")
async def upload_document(
    fault_id: int, 
    file: UploadFile = File(...), 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_roles(current_user, {"Admin", "Engineer"})
    db_fault = db.query(models.FaultLog).filter(models.FaultLog.id == fault_id).first()
    if not db_fault:
        raise HTTPException(status_code=404, detail="Fault not found")

    original_name = Path(file.filename or "").name
    extension = Path(original_name).suffix.lower()
    if extension not in ALLOWED_UPLOAD_EXTENSIONS or file.content_type not in ALLOWED_UPLOAD_TYPES:
        raise HTTPException(status_code=400, detail="Only JPG, PNG, WEBP, and PDF attachments are allowed")
    
    upload_dir = "uploads"
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir)

    safe_stem = "".join(ch for ch in Path(original_name).stem if ch.isalnum() or ch in ("-", "_")).strip() or "attachment"
    safe_filename = f"{fault_id}_{int(datetime.now().timestamp())}_{safe_stem}{extension}"
    file_path = os.path.join(upload_dir, safe_filename)

    total = 0
    with open(file_path, "wb") as buffer:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            total += len(chunk)
            if total > MAX_UPLOAD_BYTES:
                buffer.close()
                os.remove(file_path)
                raise HTTPException(status_code=413, detail="Attachment exceeds the 10MB limit")
            buffer.write(chunk)

    db_fault.attachment_path = f"uploads/{safe_filename}"
    db_fault.updated_at = datetime.now()
    _record_timeline(db, fault_id, "Attachment Added", original_name, current_user.username)
    db.commit()
    return {"filename": safe_filename, "path": db_fault.attachment_path}

@app.get("/faults/{fault_id}/comments", response_model=List[schemas.FaultCommentOut])
def list_fault_comments(
    fault_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    fault = db.query(models.FaultLog).filter(models.FaultLog.id == fault_id).first()
    if not fault:
        raise HTTPException(status_code=404, detail="Fault not found")
    return (
        db.query(models.FaultComment)
        .filter(models.FaultComment.fault_id == fault_id)
        .order_by(models.FaultComment.created_at.asc())
        .all()
    )

@app.post("/faults/{fault_id}/comments", response_model=schemas.FaultCommentOut)
async def add_fault_comment(
    fault_id: int,
    payload: schemas.FaultCommentCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    _require_roles(current_user, {"Admin", "Engineer"})
    fault = db.query(models.FaultLog).filter(models.FaultLog.id == fault_id).first()
    if not fault:
        raise HTTPException(status_code=404, detail="Fault not found")

    comment_text = (payload.comment or "").strip()
    if not comment_text:
        raise HTTPException(status_code=400, detail="comment is required")

    comment = models.FaultComment(fault_id=fault_id, comment=comment_text, created_by=current_user.username)
    db.add(comment)
    fault.updated_at = datetime.now()
    _record_timeline(db, fault_id, "Comment Added", comment_text, current_user.username)
    db.commit()
    db.refresh(comment)
    await manager.broadcast({"event": "fault_comment_added", "id": fault_id, "ticket": fault.ticket_number, "isp": fault.isp_name})
    return comment

@app.get("/faults/{fault_id}/timeline", response_model=List[schemas.FaultTimelineOut])
def list_fault_timeline(
    fault_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    fault = db.query(models.FaultLog).filter(models.FaultLog.id == fault_id).first()
    if not fault:
        raise HTTPException(status_code=404, detail="Fault not found")
    return (
        db.query(models.FaultTimeline)
        .filter(models.FaultTimeline.fault_id == fault_id)
        .order_by(models.FaultTimeline.created_at.asc(), models.FaultTimeline.id.asc())
        .all()
    )

# Reporting & Stats
@app.get("/stats/", response_model=schemas.StatsSummary)
def get_stats(db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    total = db.query(models.FaultLog).count()
    open_count = db.query(models.FaultLog).filter(models.FaultLog.status == "Open").count()
    active_count = db.query(models.FaultLog).filter(models.FaultLog.status.in_(list(ACTIVE_STATUSES))).count()
    critical_count = db.query(models.FaultLog).filter(
        models.FaultLog.status.in_(list(ACTIVE_STATUSES)),
        models.FaultLog.severity == "Critical"
    ).count()
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    week_ago = datetime.utcnow() - timedelta(days=7)
    month_ago = datetime.utcnow() - timedelta(days=30)
    resolved_today = db.query(models.FaultLog).filter(
        models.FaultLog.status.in_(["Resolved", "Closed"]),
        models.FaultLog.resolved_at >= today
    ).count()
    
    day_logs = db.query(models.FaultLog).filter(models.FaultLog.created_at >= today).count()
    week_logs = db.query(models.FaultLog).filter(models.FaultLog.created_at >= week_ago).count()
    month_logs = db.query(models.FaultLog).filter(models.FaultLog.created_at >= month_ago).count()
    
    resolution_minutes = []
    resolved_logs = db.query(models.FaultLog).filter(models.FaultLog.resolved_at != None).all()  # noqa: E711
    for log in resolved_logs:
        if log.created_at and log.resolved_at:
            minutes = (log.resolved_at - log.created_at).total_seconds() / 60
            if minutes >= 0:
                resolution_minutes.append(minutes)
    avg_res = round(sum(resolution_minutes) / len(resolution_minutes), 1) if resolution_minutes else 0.0

    most_affected = (
        db.query(models.FaultLog.isp_name, func.count(models.FaultLog.id).label("fault_count"))
        .group_by(models.FaultLog.isp_name)
        .order_by(func.count(models.FaultLog.id).desc())
        .first()
    )
    latest = db.query(models.FaultLog).order_by(models.FaultLog.created_at.desc()).first()
    
    return {
        "total_logs": total,
        "open_issues": open_count,
        "active_faults": active_count,
        "critical_incidents": critical_count,
        "resolved_today": resolved_today,
        "day_logs_count": day_logs,
        "week_logs_count": week_logs,
        "month_logs_count": month_logs,
        "avg_resolution_min": avg_res,
        "most_affected_isp": most_affected[0] if most_affected else None,
        "latest_ticket": latest.ticket_number if latest else None,
        "latest_location": latest.location if latest else None,
    }

# Export System (Enhanced)
@app.get("/export/{format}")
def export_logs(format: str, period: str = "all", db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    logs_data = read_faults(period=period, db=db, current_user=current_user)
    if format == "xlsx":
        df = pd.DataFrame([{
            "Ticket": l.ticket_number or l.id, "ID": l.id, "ISP": l.isp_name, "Location": l.location, 
            "Type": l.fault_type, "Severity": l.severity, "Status": l.status,
            "Assigned To": l.assigned_to or "",
            "Date": l.created_at.strftime("%Y-%m-%d") if l.created_at else "",
            "Time Created": l.created_at.strftime("%H:%M") if l.created_at else "",
            "Time Resolved": l.resolved_at.strftime("%H:%M") if l.resolved_at else "",
            "Resolution Note": l.resolution_note or "",
            "SLA Breach": l.is_sla_breach
        } for l in logs_data])
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Faults')
            
            # Apply styling
            workbook = writer.book
            worksheet = writer.sheets['Faults']
            
            from openpyxl.styles import PatternFill
            fill_yellow = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
            fill_green = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid') # Soft Green
            fill_red = PatternFill(start_color='FFFF6666', end_color='FFFF6666', fill_type='solid') # Soft Red to keep text visible
            
            headers = [cell.value for cell in worksheet[1]]
            col_date = headers.index("Date") + 1 if "Date" in headers else None
            col_created = headers.index("Time Created") + 1 if "Time Created" in headers else None
            col_resolved = headers.index("Time Resolved") + 1 if "Time Resolved" in headers else None
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row)):
                if row_idx == 0: continue # Skip header
                if col_date: row[col_date-1].fill = fill_yellow
                if col_created: row[col_created-1].fill = fill_green
                if col_resolved: row[col_resolved-1].fill = fill_red

        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=report.xlsx"})
    
    if format == "isp-report":
        chart_data = _build_chart_data(db)
        docx_stream = export_service.ExportService.isp_report_docx(chart_data)
        return StreamingResponse(
            docx_stream,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": "attachment; filename=ISP_Performance_Report.docx"}
        )
    
    # Fallback to existing logic for docx/pdf if needed
    raise HTTPException(status_code=400, detail="Format not fully implemented in enterprise mode yet")

# Chart Data
def _get_isp_names(db: Session) -> list[str]:
    names = [
        row[0] for row in db.query(models.FaultLog.isp_name)
        .filter(models.FaultLog.isp_name != None)  # noqa: E711
        .distinct()
        .all()
        if row[0]
    ]
    defaults = ["Powertel", "Starlink", "Liquid"]
    return sorted(set(defaults + names))

def _build_chart_data(db: Session) -> dict:
    """Compute per-ISP performance metrics from the DB."""
    all_isps = _get_isp_names(db)
    isp_stats = []
    for isp in all_isps:
        total = db.query(models.FaultLog).filter(models.FaultLog.isp_name == isp).count()
        resolved = db.query(models.FaultLog).filter(
            models.FaultLog.isp_name == isp,
            models.FaultLog.status.in_(["Resolved", "Closed"])
        ).count()
        critical = db.query(models.FaultLog).filter(
            models.FaultLog.isp_name == isp,
            models.FaultLog.severity == "Critical"
        ).count()
        resolution_rate = (resolved / total * 100) if total > 0 else 0
        isp_stats.append({
            "name": isp,
            "total_faults": total,
            "resolved": resolved,
            "critical": critical,
            "resolution_rate": round(resolution_rate, 1)
        })
    return {"isps": isp_stats}

@app.get("/chart-data/")
def get_chart_data(db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    return _build_chart_data(db)

# Speed Test — uses Cloudflare's public speed endpoint (no external library needed)
@app.get("/speedtest", response_model=schemas.SpeedTestResult)
async def run_speedtest(current_user: models.User = Depends(get_current_user)):
    """Measure real internet speed via Cloudflare's public speed-test API."""
    import urllib.request, time, random, ssl

    CLOUDFLARE = "https://speed.cloudflare.com"
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    def _measure_ping(url: str, attempts: int = 5) -> float:
        times = []
        for _ in range(attempts):
            try:
                t0 = time.perf_counter()
                req = urllib.request.Request(url + "/", method="HEAD")
                req.add_header("User-Agent", "Mozilla/5.0")
                urllib.request.urlopen(req, timeout=5, context=ctx)
                times.append((time.perf_counter() - t0) * 1000)
            except Exception:
                pass
        return round(min(times) if times else 999, 2)

    def _measure_download(bytes_count: int = 10_000_000) -> float:
        url = f"{CLOUDFLARE}/__down?bytes={bytes_count}"
        req = urllib.request.Request(url)
        req.add_header("User-Agent", "Mozilla/5.0")
        req.add_header("Accept-Encoding", "identity")
        t0 = time.perf_counter()
        with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
            data = resp.read()
        elapsed = time.perf_counter() - t0
        return round((len(data) * 8) / elapsed / 1_000_000, 2)  # Mbps

    def _measure_upload(bytes_count: int = 4_000_000) -> float:
        payload = random.randbytes(bytes_count)
        url = f"{CLOUDFLARE}/__up"
        req = urllib.request.Request(url, data=payload, method="POST")
        req.add_header("User-Agent", "Mozilla/5.0")
        req.add_header("Content-Type", "application/octet-stream")
        req.add_header("Content-Length", str(len(payload)))
        t0 = time.perf_counter()
        with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
            resp.read()
        elapsed = time.perf_counter() - t0
        return round((bytes_count * 8) / elapsed / 1_000_000, 2)  # Mbps

    def _run_all():
        print("Starting Ping measurement...")
        ping = _measure_ping(CLOUDFLARE)
        print(f"Ping: {ping} ms. Starting Download measurement...")
        download = _measure_download()
        print(f"Download: {download} Mbps. Starting Upload measurement...")
        upload = _measure_upload()
        print(f"Upload: {upload} Mbps. Speed test complete.")
        return {"ping": ping, "download": download, "upload": upload}

    loop = asyncio.get_event_loop()
    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(None, _run_all),
            timeout=60.0
        )
        return {
            "download": result["download"],
            "upload": result["upload"],
            "ping": result["ping"],
            "server": "Cloudflare Speed Test (Global CDN)",
            "timestamp": datetime.now()
        }
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="Speed test timed out. Please check your connection and try again.")
    except Exception as e:
        print(f"Speedtest failed: {e}")
        raise HTTPException(status_code=500, detail=f"Speed test error: {str(e)}")


# Rich Analytics Endpoint
@app.get("/analytics/")
def get_analytics(db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    """Rich analytics data for the dashboard charts."""
    print(f"Generating analytics for user: {current_user.username}")
    all_faults = db.query(models.FaultLog).all()
    total = len(all_faults)
    resolved = [f for f in all_faults if f.status in {"Resolved", "Closed"}]
    open_faults = [f for f in all_faults if f.status in ACTIVE_STATUSES]
    critical = [f for f in all_faults if f.severity == "Critical"]
    sla_breaches = [f for f in all_faults if f.is_sla_breach]

    # Resolution times
    resolve_times = []
    for f in resolved:
        if f.resolved_at and f.created_at:
            mins = (f.resolved_at - f.created_at).total_seconds() / 60
            if mins >= 0:
                resolve_times.append(mins)
    avg_resolve_min = round(sum(resolve_times) / len(resolve_times), 1) if resolve_times else 0

    # Daily trend — last 30 days
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    daily_trend = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        day_end = day + timedelta(days=1)
        count = sum(1 for f in all_faults if f.created_at and day <= f.created_at < day_end)
        resolved_count = sum(1 for f in resolved if f.resolved_at and day <= f.resolved_at < day_end)
        daily_trend.append({
            "date": day.strftime("%d %b"),
            "faults": count,
            "resolved": resolved_count
        })

    # By ISP
    all_isps = _get_isp_names(db)
    isp_data = []
    for isp in all_isps:
        isp_faults = [f for f in all_faults if f.isp_name == isp]
        isp_resolved = [f for f in isp_faults if f.status in {"Resolved", "Closed"}]
        isp_critical = [f for f in isp_faults if f.severity == "Critical"]
        rate = round(len(isp_resolved) / len(isp_faults) * 100, 1) if isp_faults else 0
        isp_data.append({
            "name": isp,
            "total": len(isp_faults),
            "resolved": len(isp_resolved),
            "open": len(isp_faults) - len(isp_resolved),
            "critical": len(isp_critical),
            "resolution_rate": rate
        })

    # By severity
    severities = ["Minor", "Major", "Critical"]
    severity_colors = {"Minor": "#7FBFB3", "Major": "#C6A75C", "Critical": "#C1121F"}
    by_severity = [
        {"name": s, "value": sum(1 for f in all_faults if f.severity == s), "color": severity_colors[s]}
        for s in severities
    ]

    # By fault type
    fault_types = {}
    for f in all_faults:
        if f.fault_type:
            fault_types[f.fault_type] = fault_types.get(f.fault_type, 0) + 1
    type_colors = ["#7FBFB3", "#C6A75C", "#C1121F", "#1E3A8A", "#64748b", "#a855f7"]
    by_type = sorted(
        [{"name": k, "value": v, "color": type_colors[i % len(type_colors)]} for i, (k, v) in enumerate(fault_types.items())],
        key=lambda x: x["value"], reverse=True
    )

    # By location
    locations = {}
    for f in all_faults:
        if f.location:
            locations[f.location] = locations.get(f.location, 0) + 1
    by_location = sorted(
        [{"name": k, "value": v} for k, v in locations.items()],
        key=lambda x: x["value"], reverse=True
    )[:8]

    return {
        "summary": {
            "total": total,
            "resolved": len(resolved),
            "open": len(open_faults),
            "critical": len(critical),
            "sla_breaches": len(sla_breaches),
            "resolution_rate": round(len(resolved) / total * 100, 1) if total else 0,
            "avg_resolve_min": avg_resolve_min
        },
        "daily_trend": daily_trend,
        "by_isp": isp_data,
        "by_severity": by_severity,
        "by_type": by_type[:6],
        "by_location": by_location
    }


# WebSockets
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)
